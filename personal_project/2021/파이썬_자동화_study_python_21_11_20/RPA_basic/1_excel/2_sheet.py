from openpyxl import Workbook
from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()#시트 추가.
ws.title = "MySheet"
ws.sheet_properties.tabColor="ff66ff"

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet",2)#2번째 index에 sheet 생성

new_ws = wb["NewSheet"]#Dict 형태로 sheet에 접근

print(wb.sheetnames)#모두의 sheet 이름 확인

new_ws["A1"] = "Test"
target  = wb.copy_worksheet(new_ws)
target.title = "Copide Sheet"

wb.save("create_sheet_2.xlsx")
wb.close()

# 11/20 3시 종료. 집에서 라던가 멀티 노트북에서 상황 이어서 할것
# https://www.youtube.com/watch?v=exgO1LFl9x8&t=0s
# 24:49 부터 이어서 함(5.셀 기본)